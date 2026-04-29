import os
import sys
import json
import re
import urllib.request
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import hashlib

# --- PYTHON 3.13 DRAG-AND-DROP FIX ---
if 'tkinter.tix' not in sys.modules:
    sys.modules['tkinter.tix'] = tk
    tk.tix = tk

# --- Graceful Degradation ---
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False

import customtkinter as ctk
import datetime
import webbrowser
import subprocess
import platform

# --- Excel Export Libraries ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# =====================================================================
# ⚙️ APP SETTINGS & AUTO-UPDATER CONFIG
# =====================================================================
APP_VERSION = "1.3.0"

# When you are ready to use the updater, replace this URL with your raw GitHub JSON link!
GITHUB_UPDATE_URL = "https://raw.githubusercontent.com/nzokajohn/RenamerApp/main/version.json"

SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".renamer_pro_settings.json")
ALLOWED_EXTENSIONS = {'.mp4', '.mov', '.mxf', '.wav', '.mp3', '.m4a', '.aac', '.jpg', '.jpeg', '.png', '.dpx', '.exr'}

# Set the visual theme to modern dark
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

MAIN_BG = "#1C1C1E"
CARD_BG = "#2C2C2E"
FONT_MAIN = ("Helvetica Neue", 13)
FONT_BOLD = ("Helvetica Neue", 13, "bold")
FONT_H1 = ("Helvetica Neue", 18, "bold")

# --- USER MANUAL TEXT ---
MANUAL_TEXT = r"""📖 MEDIA FILE RENAMER PRO – USER MANUAL

Welcome to Media File Renamer Pro! This application is designed to completely automate the tedious process of renaming media files for CMS delivery. It standardizes filenames, prevents accidental overwrites, and automatically generates organized Delivery Manifests.

--------------------------------------------------
🚀 QUICK START GUIDE
--------------------------------------------------
STEP 1: CHOOSE A PROFILE (OR SET METADATA)
If you have saved presets for different projects, choose one from the top dropdown! 
🤝 TEAM SHARING: You can use the Export (⬆️) and Import (⬇️) buttons to share your perfect Profiles with other members of your team!

Otherwise, set your Global Metadata manually:
• Language: Audio language code (e.g., 'en'). Leave blank if non-verbal (defaults to 'n/a').
• Category: Primary focus (e.g., 'stretching'). Spaces automatically convert to underscores.
• Aspect Ratio: Target delivery format. Check "Auto-Detect" to let the app scan video dimensions and automatically cascade the correct ratio to all stems in that bundle!
• Application: Target platform abbreviation (e.g., 'yg').

STEP 2: CHOOSE YOUR MODE
Single Folder Mode:
• Select the Content Type radio button (e.g., Complete Video, Stem 1).
• Click 'Select Folder & Rename' to execute.

Batch Processing Mode (Recommended):
• Check 'Enable Batch Processing'.
• Drag and Drop multiple folders directly into the app's drop zone.
• Auto-Detect Magic: The app reads folder names and automatically assigns the Content Type (e.g., "Audio Only" auto-selects Stem 3). You can manually adjust the dropdowns if it guesses wrong.

STEP 3: PREVIEW AND EXECUTE
• Click '👁 Preview Naming' for a dry-run to see what files will look like safely.
• Click '🚀 Execute' to process the queue.

--------------------------------------------------
🧠 ADVANCED FEATURES
--------------------------------------------------
Click '⚙️ Show Advanced Options' to unlock powerful bulk-editing tools:

1. Standard Find & Replace: Type a word to remove or change.
2. Force Casing: Enforce strict base filename casing (lowercase, UPPERCASE, Title Case).
3. Counter Padding & Sorting: Control how duplicate files in a bundle are numbered (e.g., _01 vs _001) and sort them by Alphabetical, Creation Date, or Modified Date so they number perfectly based on when they were shot!
4. Regex (Regular Expressions): Instead of exact words, use code to find patterns!
5. MD5 Checksums: Check this to mathematically calculate unique file hashes for strict CMS deliveries.
6. Custom Naming Template: Total control over the output format using bracket placeholders!

--------------------------------------------------
🛡️ SAFETY & METRICS
--------------------------------------------------
• Emoji Stripper: The app automatically detects and deletes Emojis and hidden illegal CMS characters from files on the fly.
• Pre-Flight Checker: Before renaming, if it detects a filename over 150 characters, it aborts instantly.
• Manifest Metrics: Your Excel report automatically extracts the exact Duration of your videos.
• 1-Click Quick Undo: Safely reverts all files/folders from the bottom up, and deletes the Excel report.
"""

# --- Custom Tooltip Class (Standard) ---
class ToolTip:
    def __init__(self, widget, text, wrap_length=280):
        self.widget = widget
        self.text = text
        self.wrap_length = wrap_length
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x = self.widget.winfo_rootx() + self.widget.winfo_width() + 15
        y = self.widget.winfo_rooty() - 5
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left', background="#3A3A3C", foreground="white", 
                         relief='flat', font=("Arial", 12), padx=10, pady=8, wraplength=self.wrap_length)
        label.pack()

    def leave(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

# --- Custom Rich Tooltip Class (Color Coded) ---
class RichToolTip:
    def __init__(self, widget, segments, width=45, height=13):
        self.widget = widget
        self.segments = segments
        self.width = width
        self.height = height
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x = self.widget.winfo_rootx() + self.widget.winfo_width() + 15
        y = self.widget.winfo_rooty() - 50 
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        
        text_widget = tk.Text(tw, background="#3A3A3C", foreground="white", relief="flat", 
                              font=("Helvetica Neue", 12), padx=10, pady=8, wrap="word", 
                              width=self.width, height=self.height, highlightthickness=0)
        text_widget.pack()
        
        for txt, color in self.segments:
            tag_name = f"color_{color.replace('#', '')}"
            text_widget.tag_configure(tag_name, foreground=color)
            text_widget.insert("end", txt, tag_name)
            
        text_widget.configure(state="disabled")

    def leave(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

# --- Global Logic Functions ---
def get_creation_time(path):
    stat = os.stat(path)
    try: return stat.st_birthtime
    except AttributeError: return stat.st_ctime

def get_media_duration(filepath):
    if platform.system() != "Darwin": return "N/A"
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in {'.mp4', '.mov', '.mxf', '.wav', '.mp3', '.m4a', '.aac'}: return "N/A"
    
    try:
        cmd = ['mdls', '-name', 'kMDItemDurationSeconds', filepath]
        res = subprocess.run(cmd, capture_output=True, text=True, timeout=2)
        match = re.search(r'kMDItemDurationSeconds\s*=\s*([\d\.]+)', res.stdout)
        if match:
            seconds = float(match.group(1))
            if seconds == 0: return "N/A"
            td = datetime.timedelta(seconds=int(seconds))
            return str(td)
    except: pass
    return "N/A"

def calculate_md5(filepath):
    hash_md5 = hashlib.md5()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except Exception:
        return "Error"

def get_auto_aspect_ratio(filepath, default_aspect):
    if platform.system() != "Darwin":
        return default_aspect
    ext = os.path.splitext(filepath)[1].lower()
    if ext in {'.wav', '.mp3', '.m4a', '.aac'}:
        return default_aspect
    try:
        cmd = ['mdls', '-name', 'kMDItemPixelWidth', '-name', 'kMDItemPixelHeight', filepath]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=2)
        w_match = re.search(r'kMDItemPixelWidth\s*=\s*(\d+)', result.stdout)
        h_match = re.search(r'kMDItemPixelHeight\s*=\s*(\d+)', result.stdout)
        if w_match and h_match:
            w = float(w_match.group(1))
            h = float(h_match.group(1))
            if w == 0 or h == 0: return default_aspect
            ratio = w / h
            if ratio >= 1.2: return "16x9"
            elif ratio <= 0.8: return "9x16"
            else: return "1x1"
    except Exception:
        pass
    return default_aspect

def move_to_trash(filepath):
    if platform.system() == "Darwin":
        try:
            abs_path = os.path.abspath(filepath)
            script = f'tell application "Finder" to delete POSIX file "{abs_path}"'
            subprocess.run(['osascript', '-e', script], check=True, capture_output=True)
        except: os.remove(filepath)
    else: os.remove(filepath)

def process_revert_from_log(log_file):
    folder = os.path.dirname(log_file)
    success_count = 0
    try:
        with open(log_file, 'r') as f: lines = f.readlines()
        file_reverts, dir_reverts = [], []
        
        for line in lines[1:]:
            parts = line.strip().split('\t')
            if len(parts) == 3 and parts[0] == 'DIR_RENAME':
                dir_reverts.append((parts[1], parts[2])) 
            elif len(parts) == 2:
                file_reverts.append((parts[0], parts[1])) 
                
        for orig, new in file_reverts:
            current_path = os.path.join(folder, new)
            original_path = os.path.join(folder, orig)
            if os.path.exists(current_path):
                os.rename(current_path, original_path)
                success_count += 1
                
        for orig, new in dir_reverts:
            current_path = os.path.join(folder, new)
            original_path = os.path.join(folder, orig)
            if os.path.exists(current_path):
                os.rename(current_path, original_path)
                success_count += 1
        
        if success_count > 0 or len(lines) <= 1: move_to_trash(log_file)
        return success_count
    except Exception: return 0

def apply_find_replace(name, find_str, replace_str, use_regex, smart_prefix):
    if not find_str: return name
    if smart_prefix:
        pattern = r'(?<![A-Za-z])' + re.escape(find_str) + r'(\s*\d+)'
        try: return re.sub(pattern, replace_str + r'\g<1>', name)
        except: return name
    elif use_regex:
        try: return re.sub(find_str, replace_str, name)
        except: return name
    else:
        return name.replace(find_str, replace_str)

def generate_safe_name(folder_path, filename, stem_val, lang, category, aspect, app, find_str, replace_str, use_regex, smart_prefix, template_str, case_format, pad_len):
    name_id, ext = os.path.splitext(filename)
    if ext.lower() not in ALLOWED_EXTENSIONS: return None, None

    stem_map = {"Complete": "", "Video": "_VIDEO", "Speech": "_SPEECH", "Music": "_MUSIC"}
    stem_suffix = stem_map.get(stem_val, "")
    
    name_id = apply_find_replace(name_id, find_str, replace_str, use_regex, smart_prefix)
    
    if case_format == "lowercase":
        name_id = name_id.lower()
    elif case_format == "UPPERCASE":
        name_id = name_id.upper()
    elif case_format == "Title Case":
        name_id = name_id.title()
        
    name_id = name_id.replace(" ", "_")
    
    try:
        base_new_name = template_str.format(
            name=name_id, stem=stem_suffix, lang=lang,
            cat=category, aspect=aspect, app=app
        )
    except Exception:
        base_new_name = f"{name_id}{stem_suffix}+lang={lang}&category={category}&aspect={aspect}&app={app}&master"
    
    # --- AUTO-STRIP EMOJIS & ILLEGAL CHARACTERS ---
    base_new_name = "".join(c for c in base_new_name if ord(c) < 128)
    
    final_name = f"{base_new_name}{ext}"
    counter = 1
    while os.path.exists(os.path.join(folder_path, final_name)):
        final_name = f"{base_new_name}_{str(counter).zfill(pad_len)}{ext}"
        counter += 1
        
    return final_name, name_id

def workout_bundle_sort_key(row):
    generated_name = row[2]
    pure_name_id = row[3] 
    pure_base = re.sub(r'[^A-Za-z0-9]', '', pure_name_id).lower()
    group_key = [int(text) if text.isdigit() else text for text in re.split(r'(\d+)', pure_base)]
    
    if "_VIDEO" in generated_name: content_rank = 1
    elif "_SPEECH" in generated_name: content_rank = 2
    elif "_MUSIC" in generated_name: content_rank = 3
    else: content_rank = 0 
        
    return (group_key, content_rank)

# --- Dynamic App Base ---
if DND_AVAILABLE:
    class AppBase(ctk.CTk, TkinterDnD.DnDWrapper):
        def __init__(self):
            super().__init__()
            self.TkdndVersion = TkinterDnD._require(self)
else:
    class AppBase(ctk.CTk):
        pass

# --- Main App Class ---
class RenamerApp(AppBase):
    def __init__(self):
        super().__init__()
        
        self.title(f"Media File Renamer Pro - v{APP_VERSION}")
        self.geometry("560x840") 
        self.minsize(540, 650) 
        self.configure(fg_color=MAIN_BG, padx=20, pady=20)
        self.protocol("WM_DELETE_WINDOW", self.save_settings_and_close)

        # Settings & Presets
        self.presets = {"Default": {}}
        self.preset_var = tk.StringVar(value="Default")

        # Variables
        self.lang_var = tk.StringVar(value="en")
        self.category_var = tk.StringVar(value="stretching")
        self.aspect_var = tk.StringVar(value="16x9")
        self.auto_aspect_var = tk.BooleanVar(value=True) 
        self.app_var = tk.StringVar(value="yg")
        
        self.find_var = tk.StringVar(value="")
        self.replace_var = tk.StringVar(value="")
        self.regex_var = tk.BooleanVar(value=False)
        self.smart_prefix_var = tk.BooleanVar(value=False)
        self.rename_folders_var = tk.BooleanVar(value=False)
        self.case_var = tk.StringVar(value="None") 
        self.padding_var = tk.StringVar(value="01")
        self.sort_var = tk.StringVar(value="Alphabetical")
        self.md5_var = tk.BooleanVar(value=False)
        
        self.batch_mode_var = tk.BooleanVar(value=False)
        self.stem_var = tk.StringVar(value="Complete")
        
        self.batch_folders = []
        self.last_generated_logs = []
        self.last_generated_excel = "" 
        self.manual_window = None

        self.setup_ui()
        self.load_settings()
        
        threading.Thread(target=self.check_for_updates, daemon=True).start()

    def check_for_updates(self):
        try:
            import ssl
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE

            req = urllib.request.Request(GITHUB_UPDATE_URL, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5, context=ctx) as response:
                data = json.loads(response.read().decode())
                latest_version = data.get("version", APP_VERSION)
                
                # --- SMART VERSION COMPARISON ---
                current_v = [int(x) if x.isdigit() else 0 for x in APP_VERSION.split('.')]
                latest_v = [int(x) if x.isdigit() else 0 for x in latest_version.split('.')]
                
                if latest_v > current_v:
                    self.after(1000, lambda: self.prompt_update(data))
        except Exception as e: 
            print(f"Updater check bypassed or failed: {e}")

    def prompt_update(self, data):
        version = data.get("version")
        notes = data.get("release_notes", "No release notes provided.")
        url = data.get("download_url", "")
        msg = f"A new version (v{version}) of Media File Renamer Pro is available!\n\nRelease Notes:\n{notes}\n\nWould you like to download it now?"
        if messagebox.askyesno("Update Available", msg):
            if url: webbrowser.open(url)

    def setup_ui(self):
        # =================================================================
        # PINNED BOTTOM FOOTER & ACTIONS
        # =================================================================
        self.footer_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.footer_frame.pack(side="bottom", fill="x", pady=(0, 0))
        link_label = ctk.CTkLabel(self.footer_frame, text="nzokajohn", font=("Arial", 12, "underline"), text_color="#0A84FF", cursor="hand2")
        link_label.pack()
        link_label.bind("<Button-1>", lambda e: webbrowser.open("https://www.linkedin.com/in/nzokajohn/"))

        self.status_lbl = ctk.CTkLabel(self, text="Ready", font=("Arial", 12), text_color="#8E8E93")
        self.status_lbl.pack(side="bottom", pady=(2, 2))

        action_group = ctk.CTkFrame(self, fg_color="transparent")
        action_group.pack(side="bottom", fill="x", pady=(5, 0))
        
        btn_frame = ctk.CTkFrame(action_group, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 5))
        btn_frame.columnconfigure((0,1), weight=1)
        
        ctk.CTkButton(btn_frame, text="👁 Preview Naming", fg_color="#3A3A3C", hover_color="#48484A", font=FONT_BOLD, height=36, command=self.preview_rename).grid(row=0, column=0, padx=(0,5), sticky="we")
        self.execute_btn = ctk.CTkButton(btn_frame, text="🚀 Select Folder & Rename", fg_color="#0A84FF", hover_color="#0066CC", font=FONT_BOLD, height=36, command=self.execute_rename)
        self.execute_btn.grid(row=0, column=1, padx=(5,0), sticky="we")

        ctk.CTkButton(action_group, text="↩ Undo Latest Renaming", fg_color="transparent", text_color="#E5E5EA", border_color="#48484A", border_width=1, hover_color="#2C2C2E", font=FONT_MAIN, height=30, command=self.execute_revert).pack(fill="x", pady=2)
        ctk.CTkButton(action_group, text="🗑 Deep Sweep Backup Logs", fg_color="transparent", text_color="#FF453A", hover_color="#3A1D1D", font=FONT_MAIN, height=30, command=self.execute_delete_logs).pack(fill="x", pady=2)

        # =================================================================
        # SCROLLABLE MAIN CONTENT (Tighter Padding)
        # =================================================================
        self.main_scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.main_scroll.pack(side="top", fill="both", expand=True)

        # Header
        header_frame = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 2))
        ctk.CTkLabel(header_frame, text="Metadata Standardization", font=FONT_H1).pack(side="left")
        manual_btn = ctk.CTkButton(header_frame, text="📖 User Manual", font=FONT_BOLD, fg_color=CARD_BG, text_color="white", hover_color="#3A3A3C", width=120, height=30, command=self.open_manual)
        manual_btn.pack(side="right")

        # --- PRESET MANAGER ROW ---
        preset_frame = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        preset_frame.pack(fill="x", pady=(2, 5))
        
        ctk.CTkLabel(preset_frame, text="Profile:", font=FONT_BOLD).pack(side="left")
        self.preset_cb = ctk.CTkComboBox(preset_frame, variable=self.preset_var, values=list(self.presets.keys()), command=self.load_preset_ui, font=FONT_MAIN, border_width=1, fg_color="#1C1C1E", button_color="#3A3A3C", width=130)
        self.preset_cb.pack(side="left", padx=(10, 5))
        
        ctk.CTkButton(preset_frame, text="💾 Save", font=FONT_BOLD, width=50, height=28, fg_color="#3A3A3C", hover_color="#48484A", command=self.save_preset).pack(side="left", padx=2)
        ctk.CTkButton(preset_frame, text="🗑", font=FONT_BOLD, width=35, height=28, fg_color="#3A3A3C", hover_color="#FF453A", command=self.delete_preset).pack(side="left", padx=2)
        
        btn_export = ctk.CTkButton(preset_frame, text="⬆️", font=FONT_BOLD, width=35, height=28, fg_color="#3A3A3C", hover_color="#48484A", command=self.export_preset)
        btn_export.pack(side="left", padx=2)
        ToolTip(btn_export, "Export this Profile to share with a team member.")
        
        btn_import = ctk.CTkButton(preset_frame, text="⬇️", font=FONT_BOLD, width=35, height=28, fg_color="#3A3A3C", hover_color="#48484A", command=self.import_preset)
        btn_import.pack(side="left", padx=(2, 0))
        ToolTip(btn_import, "Import a Profile shared by a team member.")

        # --- CARD 1: Core Inputs ---
        input_card = ctk.CTkFrame(self.main_scroll, fg_color=CARD_BG, corner_radius=10)
        input_card.pack(fill="x", pady=(0, 5), ipadx=8, ipady=8)
        input_card.columnconfigure(1, weight=1)

        fields = [
            ("Language", self.lang_var, "Audio language code (e.g., 'en'). Leave blank for 'n/a'."),
            ("Category", self.category_var, "Primary focus. Spaces automatically become underscores."),
            ("Application", self.app_var, "Platform abbreviations (e.g., 'yg, mb').")
        ]

        for i, (label, var, tip) in enumerate(fields):
            lbl = ctk.CTkLabel(input_card, text=f"{label}:", font=FONT_BOLD)
            lbl.grid(row=i, column=0, sticky="w", pady=5, padx=(10, 15))
            entry = ctk.CTkEntry(input_card, textvariable=var, font=FONT_MAIN, border_width=1, fg_color="#1C1C1E")
            entry.grid(row=i, column=1, sticky="we", pady=5, padx=(0, 10))
            ToolTip(lbl, tip); ToolTip(entry, tip)

        lbl_aspect = ctk.CTkLabel(input_card, text="Aspect Ratio:", font=FONT_BOLD)
        lbl_aspect.grid(row=3, column=0, sticky="w", pady=5, padx=(10, 15))
        
        asp_frame = ctk.CTkFrame(input_card, fg_color="transparent")
        asp_frame.grid(row=3, column=1, sticky="we", pady=5, padx=(0, 10))
        
        self.aspect_cb = ctk.CTkComboBox(asp_frame, values=["16x9", "9x16", "1x1"], variable=self.aspect_var, state="readonly", font=FONT_MAIN, border_width=1, fg_color="#1C1C1E", button_color="#3A3A3C", width=90)
        self.aspect_cb.pack(side="left", padx=(0, 10))
        
        self.auto_aspect_chk = ctk.CTkCheckBox(asp_frame, text="Auto-Detect (Mac)", variable=self.auto_aspect_var, font=FONT_MAIN)
        self.auto_aspect_chk.pack(side="left")
        ToolTip(self.auto_aspect_chk, "Dynamically detects video dimensions. Automatically applies the correct ratio to the Complete video AND its stems!")

        # --- Batch Controls ---
        batch_row = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        batch_row.pack(fill="x", pady=(5, 0))
        self.batch_chk = ctk.CTkCheckBox(batch_row, text="Enable Batch Processing", variable=self.batch_mode_var, command=self.toggle_batch_mode, font=FONT_BOLD)
        self.batch_chk.pack(side="left")
        
        self.progress_bar = ctk.CTkProgressBar(batch_row, width=150, height=10, progress_color="#0A84FF")
        self.progress_bar.set(0)

        # --- Mode Container ---
        self.mode_container = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        self.mode_container.pack(fill="x", pady=0)
        
        # --- CARD 2: Single Mode Layout ---
        self.single_stem_frame = ctk.CTkFrame(self.mode_container, fg_color=CARD_BG, corner_radius=10)
        ctk.CTkLabel(self.single_stem_frame, text="Content Type (Folder Contents):", font=FONT_BOLD).pack(anchor="w", padx=15, pady=(5, 2))
        stems = [("Complete Video", "Complete"), ("Stem 1: Video Only", "Video"), ("Stem 2: Speech Only", "Speech"), ("Stem 3: Music Only", "Music")]
        for text, val in stems:
            ctk.CTkRadioButton(self.single_stem_frame, text=text, variable=self.stem_var, value=val, font=FONT_MAIN, radiobutton_width=18, radiobutton_height=18).pack(anchor="w", padx=25, pady=4)

        # --- CARD 3: Batch Mode Layout ---
        self.batch_frame = ctk.CTkFrame(self.mode_container, fg_color=CARD_BG, corner_radius=10)
        self.drop_zone = ctk.CTkCanvas(self.batch_frame, height=65, bg=CARD_BG, highlightthickness=0, cursor="hand2")
        self.drop_zone.pack(fill="x", pady=(10, 5), padx=15)
        self.drop_zone.bind("<Configure>", self.draw_drop_zone)
        self.drop_zone.bind("<Button-1>", lambda e: self.add_folder_manual()) 
        self.scroll_list = ctk.CTkScrollableFrame(self.batch_frame, height=180, fg_color="transparent")
        self.scroll_list.pack(fill="x", padx=5, pady=(0, 10))

        # Advanced Toggle
        self.adv_btn = ctk.CTkButton(self.main_scroll, text="⚙️ Show Advanced Options", fg_color="transparent", text_color="#0A84FF", font=FONT_BOLD, anchor="w", command=self.toggle_advanced, hover_color=MAIN_BG)
        self.adv_btn.pack(fill="x", pady=(5, 5))

        # --- CARD 4: Advanced Frame ---
        self.adv_container = ctk.CTkFrame(self.main_scroll, fg_color="transparent")
        self.adv_frame = ctk.CTkFrame(self.adv_container, fg_color=CARD_BG, corner_radius=10)
        
        # Row 0: Find & Replace
        lbl_find = tk.Label(self.adv_frame, text="Find:", bg=CARD_BG, fg="white", font=FONT_BOLD)
        lbl_find.grid(row=0, column=0, sticky="w", pady=(15, 5), padx=(15, 5))
        entry_find = ctk.CTkEntry(self.adv_frame, textvariable=self.find_var, width=120, border_width=1, fg_color="#1C1C1E")
        entry_find.grid(row=0, column=1, pady=(15, 5), sticky="w")
        ToolTip(lbl_find, "Text or pattern to search for in the original filename.")
        ToolTip(entry_find, "Text or pattern to search for in the original filename.")

        lbl_repl = tk.Label(self.adv_frame, text="Replace:", bg=CARD_BG, fg="white", font=FONT_BOLD)
        lbl_repl.grid(row=0, column=2, sticky="w", pady=(15, 5), padx=(15, 5))
        entry_repl = ctk.CTkEntry(self.adv_frame, textvariable=self.replace_var, width=120, border_width=1, fg_color="#1C1C1E")
        entry_repl.grid(row=0, column=3, pady=(15, 5), sticky="w")
        ToolTip(lbl_repl, "Text to replace the found pattern with. Leave blank to delete.")
        ToolTip(entry_repl, "Text to replace the found pattern with. Leave blank to delete.")
        
        # Row 1: Checkboxes
        chk_frame = ctk.CTkFrame(self.adv_frame, fg_color="transparent")
        chk_frame.grid(row=1, column=0, columnspan=4, sticky="w", padx=15, pady=(5, 10))
        chk_regex = ctk.CTkCheckBox(chk_frame, text="Regex", variable=self.regex_var, font=FONT_MAIN)
        chk_regex.pack(side="left", padx=(0, 15))
        ToolTip(chk_regex, "Enable Regular Expressions for advanced pattern matching (e.g., \\d+ for numbers).")

        chk_smart = ctk.CTkCheckBox(chk_frame, text="Smart Prefix", variable=self.smart_prefix_var, font=FONT_MAIN)
        chk_smart.pack(side="left", padx=(0, 15))
        ToolTip(chk_smart, "Safely change prefixes without altering the rest of the name (e.g., changes 'L1' to 'WO1', but ignores 'Full').")

        chk_folders = ctk.CTkCheckBox(chk_frame, text="Rename Subfolders", variable=self.rename_folders_var, font=FONT_MAIN)
        chk_folders.pack(side="left")
        ToolTip(chk_folders, "Apply the Find & Replace rules to the names of the folders themselves, not just the files inside.")

        # Row 2: Strict Case Formatting
        tk.Label(self.adv_frame, text="Force Casing:", bg=CARD_BG, fg="white", font=FONT_BOLD).grid(row=2, column=0, sticky="w", pady=(0, 10), padx=(15, 5))
        self.case_cb = ctk.CTkComboBox(self.adv_frame, values=["None", "lowercase", "UPPERCASE", "Title Case"], variable=self.case_var, state="readonly", font=FONT_MAIN, border_width=1, fg_color="#1C1C1E", button_color="#3A3A3C", width=140)
        self.case_cb.grid(row=2, column=1, columnspan=2, pady=(0, 10), sticky="w")
        ToolTip(self.case_cb, "Automatically forces the base filename to be strictly lowercase, UPPERCASE, or Title Case before applying templates.")

        # Row 3: Sequence Padding & Sort By
        tk.Label(self.adv_frame, text="Padding:", bg=CARD_BG, fg="white", font=FONT_BOLD).grid(row=3, column=0, sticky="w", pady=(0, 10), padx=(15, 5))
        self.pad_cb = ctk.CTkComboBox(self.adv_frame, values=["1", "01", "001", "0001"], variable=self.padding_var, state="readonly", font=FONT_MAIN, border_width=1, fg_color="#1C1C1E", button_color="#3A3A3C", width=80)
        self.pad_cb.grid(row=3, column=1, pady=(0, 10), sticky="w")
        ToolTip(self.pad_cb, "Controls how duplicate files are numbered (e.g., _1 vs _01 vs _001)")
        
        tk.Label(self.adv_frame, text="Sort By:", bg=CARD_BG, fg="white", font=FONT_BOLD).grid(row=3, column=2, sticky="w", pady=(0, 10), padx=(15, 5))
        self.sort_cb = ctk.CTkComboBox(self.adv_frame, values=["Alphabetical", "Creation Date", "Modified Date"], variable=self.sort_var, state="readonly", font=FONT_MAIN, border_width=1, fg_color="#1C1C1E", button_color="#3A3A3C", width=130)
        self.sort_cb.grid(row=3, column=3, pady=(0, 10), sticky="w")
        ToolTip(self.sort_cb, "Determines the numbering order of files inside a bundle.")

        # Row 4: MD5 Checksum
        chk_frame2 = ctk.CTkFrame(self.adv_frame, fg_color="transparent")
        chk_frame2.grid(row=4, column=0, columnspan=4, sticky="w", padx=15, pady=(0, 10))
        ctk.CTkCheckBox(chk_frame2, text="Generate MD5 Checksums (Export only - Takes longer)", variable=self.md5_var, font=FONT_MAIN).pack(side="left")

        # Row 5: Template
        tk.Label(self.adv_frame, text="Template:", bg=CARD_BG, fg="white", font=FONT_BOLD).grid(row=5, column=0, sticky="w", pady=(0, 15), padx=(15, 5))
        self.template_textbox = ctk.CTkTextbox(self.adv_frame, width=380, height=32, font=FONT_MAIN, wrap="none", border_width=1, fg_color="#1C1C1E")
        self.template_textbox.grid(row=5, column=1, columnspan=3, pady=(0, 15), sticky="w")
        self.template_textbox.tag_config("blue_var", foreground="#0A84FF")
        
        self.template_textbox.insert("1.0", "{name}{stem}+lang={lang}&category={cat}&aspect={aspect}&app={app}&master")
        self.template_textbox.bind("<KeyRelease>", self.highlight_template)
        self.template_textbox.bind("<Return>", lambda e: "break")
        
        template_tooltip_segments = [
            ("Customize your output format!\n\n", "white"),
            ("🟢 WHAT YOU CAN EDIT:\n", "#34C759"),
            ("You can freely type any symbols (like +, &, _, -) or static text (like 'master').\n\n", "white"),
            ("🔴 WHAT YOU CANNOT EDIT:\n", "#FF453A"),
            ("You must keep the exact spelling of the variables inside the brackets:\n", "white"),
            ("{name}", "#0A84FF"), (" = Base Filename\n", "white"),
            ("{stem}", "#0A84FF"), (" = Content Type Suffix (_VIDEO, etc.)\n", "white"),
            ("{lang}", "#0A84FF"), (" = Language\n", "white"),
            ("{cat}", "#0A84FF"), (" = Category\n", "white"),
            ("{aspect}", "#0A84FF"), (" = Aspect Ratio\n", "white"),
            ("{app}", "#0A84FF"), (" = Application", "white")
        ]
        RichToolTip(self.template_textbox, template_tooltip_segments)

        if DND_AVAILABLE:
            self.drop_target_register(DND_FILES)
            self.dnd_bind('<<Drop>>', self.handle_drop)
            
        self.highlight_template() 

    # --- Preset Management Logic ---
    def save_preset(self):
        dialog = ctk.CTkInputDialog(text="Enter a name for this Profile:", title="Save Profile")
        name = dialog.get_input()
        if not name: return
        name = name.strip()
        if not name: return
        
        self.presets[name] = {
            'l': self.lang_var.get(), 'c': self.category_var.get(), 'a': self.app_var.get(),
            'asp': self.aspect_var.get(), 'auto_asp': self.auto_aspect_var.get(),
            'find': self.find_var.get(), 'repl': self.replace_var.get(),
            'regex': self.regex_var.get(), 'smart': self.smart_prefix_var.get(),
            'rn_fld': self.rename_folders_var.get(), 'case': self.case_var.get(),
            'pad': self.padding_var.get(), 'sort': self.sort_var.get(), 'md5': self.md5_var.get(),
            'tmpl': self.template_textbox.get("1.0", "end-1c").strip()
        }
        
        self.preset_cb.configure(values=list(self.presets.keys()))
        self.preset_var.set(name)
        self.set_status(f"Profile '{name}' saved successfully!", "#34C759")
        
    def load_preset_ui(self, choice):
        if choice in self.presets:
            p = self.presets[choice]
            self.lang_var.set(p.get('l', 'en'))
            self.category_var.set(p.get('c', 'stretching'))
            self.app_var.set(p.get('a', 'yg'))
            self.aspect_var.set(p.get('asp', '16x9'))
            self.auto_aspect_var.set(p.get('auto_asp', True))
            self.find_var.set(p.get('find', ''))
            self.replace_var.set(p.get('repl', ''))
            self.regex_var.set(p.get('regex', False))
            self.smart_prefix_var.set(p.get('smart', False))
            self.rename_folders_var.set(p.get('rn_fld', False))
            self.case_var.set(p.get('case', 'None'))
            self.padding_var.set(p.get('pad', '01'))
            self.sort_var.set(p.get('sort', 'Alphabetical'))
            self.md5_var.set(p.get('md5', False))
            
            tmpl_val = p.get('tmpl', '')
            if not tmpl_val:
                tmpl_val = "{name}{stem}+lang={lang}&category={cat}&aspect={aspect}&app={app}&master"
                
            self.template_textbox.delete("1.0", "end")
            self.template_textbox.insert("1.0", tmpl_val)
            self.highlight_template()
            self.set_status(f"Profile '{choice}' loaded.", "#0A84FF")

    def delete_preset(self):
        name = self.preset_var.get()
        if name == "Default":
            return messagebox.showwarning("Cannot Delete", "You cannot delete the Default profile.")
        if name in self.presets:
            if messagebox.askyesno("Delete Profile", f"Are you sure you want to permanently delete the profile '{name}'?"):
                del self.presets[name]
                self.preset_cb.configure(values=list(self.presets.keys()))
                self.preset_var.set("Default")
                self.load_preset_ui("Default")
                self.set_status(f"Profile '{name}' deleted.", "#8E8E93")

    def export_preset(self):
        name = self.preset_var.get()
        if name not in self.presets or name == "Default":
            return messagebox.showwarning("Export Failed", "Please save your profile first or select a valid custom profile to export.")
        
        filepath = filedialog.asksaveasfilename(defaultextension=".json", initialfile=f"Profile_{name}.json", title="Export Profile", filetypes=[("JSON Files", "*.json")])
        if filepath:
            try:
                with open(filepath, 'w') as f:
                    json.dump(self.presets[name], f)
                self.set_status(f"Exported '{name}' successfully.", "#34C759")
            except Exception as e:
                messagebox.showerror("Export Error", f"Could not save file: {str(e)}")

    def import_preset(self):
        filepath = filedialog.askopenfilename(title="Import Profile", filetypes=[("JSON Files", "*.json")])
        if filepath:
            try:
                with open(filepath, 'r') as f:
                    data = json.load(f)
                
                dialog = ctk.CTkInputDialog(text="Name this imported profile:", title="Import Profile")
                name = dialog.get_input()
                if name:
                    name = name.strip()
                    self.presets[name] = data
                    self.preset_cb.configure(values=list(self.presets.keys()))
                    self.preset_var.set(name)
                    self.load_preset_ui(name)
                    self.set_status(f"Profile '{name}' imported successfully!", "#34C759")
            except Exception as e:
                messagebox.showerror("Import Error", f"Could not read profile file: {str(e)}")

    # --- Feature Logic ---
    def highlight_template(self, event=None):
        text = self.template_textbox.get("1.0", "end-1c")
        self.template_textbox.tag_remove("blue_var", "1.0", "end")
        valid_vars = {"{name}", "{stem}", "{lang}", "{cat}", "{aspect}", "{app}"}
        for match in re.finditer(r'\{[^{}]*\}', text):
            if match.group(0) in valid_vars:
                start_idx = f"1.0 + {match.start()} chars"
                end_idx = f"1.0 + {match.end()} chars"
                self.template_textbox.tag_add("blue_var", start_idx, end_idx)

    def open_manual(self):
        if self.manual_window and self.manual_window.winfo_exists():
            self.manual_window.focus()
            return
            
        self.manual_window = ctk.CTkToplevel(self)
        self.manual_window.title(f"User Manual - Media File Renamer Pro v{APP_VERSION}")
        self.manual_window.geometry("700x800")
        self.manual_window.minsize(550, 600)
        self.manual_window.attributes("-topmost", True)
        self.manual_window.after(200, lambda: self.manual_window.attributes("-topmost", False))
        
        textbox = ctk.CTkTextbox(self.manual_window, font=("Helvetica Neue", 14), wrap="word", fg_color=MAIN_BG, text_color="#E5E5EA")
        textbox.pack(fill="both", expand=True, padx=20, pady=(20, 10))
        textbox.insert("0.0", MANUAL_TEXT)
        textbox.configure(state="disabled") 
        
        support_btn = ctk.CTkButton(self.manual_window, text="✉️ Email Support", font=FONT_BOLD, fg_color="#0A84FF", hover_color="#0066CC", height=40, command=lambda: webbrowser.open("mailto:nthakyojohn@gmail.com?subject=Media File Renamer Pro - Support Request"))
        support_btn.pack(fill="x", padx=20, pady=(0, 20))

    def set_status(self, text, color="#8E8E93"):
        self.status_lbl.configure(text=text, text_color=color)
        self.update()

    def draw_drop_zone(self, event):
        self.drop_zone.delete("all")
        w, h = event.width, event.height
        self.drop_zone.create_rectangle(5, 5, w-5, h-5, dash=(6, 4), outline="#555555", width=2)
        text = "📂 Drag Folders Here (or Click)" if DND_AVAILABLE else "📂 Click Here to Add Folders"
        self.drop_zone.create_text(w/2, h/2, text=text, fill="#8E8E93", font=("Arial", 14))

    def toggle_batch_mode(self):
        if self.batch_mode_var.get():
            self.single_stem_frame.pack_forget()
            self.batch_frame.pack(fill="both", expand=True)
            self.execute_btn.configure(text="🚀 Execute Batch Rename")
        else:
            self.batch_frame.pack_forget()
            self.single_stem_frame.pack(fill="x")
            self.execute_btn.configure(text="🚀 Select Folder & Rename")

    def toggle_advanced(self):
        if self.adv_container.winfo_ismapped():
            self.adv_frame.pack_forget()
            self.adv_container.pack_forget()
            self.adv_btn.configure(text="⚙️ Show Advanced Options")
        else:
            self.adv_container.pack(fill="x", pady=5)
            self.adv_frame.pack(fill="x")
            self.adv_btn.configure(text="⚙️ Hide Advanced Options")
            
            # Auto-scroll to show the newly opened options
            def do_scroll():
                self.update_idletasks()
                try:
                    self.main_scroll._parent_canvas.yview_moveto(1.0)
                except Exception:
                    pass
            self.after(50, do_scroll)

    def handle_drop(self, event):
        self.batch_mode_var.set(True)
        self.toggle_batch_mode()
        paths = self.splitlist(event.data)
        for p in paths:
            if os.path.isdir(p): self.add_to_batch(p)

    def add_folder_manual(self):
        folder = filedialog.askdirectory(title="Select Folder to Add")
        if folder: self.add_to_batch(folder)

    def add_to_batch(self, path):
        row = ctk.CTkFrame(self.scroll_list, fg_color="transparent")
        row.pack(fill="x", pady=3)
        name = os.path.basename(path)
        lbl = ctk.CTkLabel(row, text=f".../{name}", width=120, anchor="w", font=FONT_MAIN)
        lbl.pack(side="left", padx=5)
        ToolTip(lbl, path, wrap_length=600)

        folder_name_lower = name.lower()
        if "speech" in folder_name_lower or "voice" in folder_name_lower: default_display, default_val = "Stem 2: Speech Only", "Speech"
        elif "music" in folder_name_lower or "audio" in folder_name_lower: default_display, default_val = "Stem 3: Music Only", "Music"
        elif "video" in folder_name_lower and "complete" not in folder_name_lower: default_display, default_val = "Stem 1: Video Only", "Video"
        else: default_display, default_val = "Complete Video", "Complete"

        s_var = tk.StringVar(value=default_val)
        opts = {"Complete Video": "Complete", "Stem 1: Video Only": "Video", "Stem 2: Speech Only": "Speech", "Stem 3: Music Only": "Music"}
        
        cb = ctk.CTkComboBox(row, values=list(opts.keys()), command=lambda v: s_var.set(opts[v]), width=170, font=FONT_MAIN, button_color="#3A3A3C", border_width=1)
        cb.set(default_display)
        cb.pack(side="left", padx=5)
        
        def remove():
            row.destroy()
            self.batch_folders = [f for f in self.batch_folders if f['path'] != path]

        ctk.CTkButton(row, text="✕", width=30, fg_color="#3A3A3C", hover_color="#FF453A", command=remove).pack(side="right", padx=(5,0))
        self.batch_folders.append({'path': path, 'stem_var': s_var})

    def format_inputs(self):
        lang = self.lang_var.get().strip() or "n/a"
        cat = self.category_var.get().strip().replace(" ", "_")
        aspect = self.aspect_var.get().strip()
        app = self.app_var.get().strip().replace(" ", "").replace(",", "_")
        find = self.find_var.get()
        repl = self.replace_var.get()
        smart_prefix = self.smart_prefix_var.get()
        rename_folders = self.rename_folders_var.get()
        template = self.template_textbox.get("1.0", "end-1c").strip()
        auto_asp = self.auto_aspect_var.get()
        case_format = self.case_var.get()
        pad_str = self.padding_var.get()
        return lang, cat, aspect, app, find, repl, smart_prefix, rename_folders, template, auto_asp, case_format, pad_str

    def preview_rename(self):
        lang, cat, aspect, app, find, repl, smart_prefix, _, template, auto_asp, case_format, pad_str = self.format_inputs()
        if not all([cat, aspect, app, template]):
            self.set_status("Missing required info for preview.", "#FF453A")
            return messagebox.showwarning("Missing Info", "Please fill out Category, Aspect Ratio, Application, and Naming Template.")

        stem = self.stem_var.get()
        if self.batch_mode_var.get() and self.batch_folders:
            stem = self.batch_folders[0]['stem_var'].get()

        pad_len = len(pad_str)
        dummy_orig = f"Project_{find}1_File.mp4" if find else "Example_Workout_File.mp4"
        dummy_new, _ = generate_safe_name(".", dummy_orig, stem, lang, cat, aspect, app, find, repl, self.regex_var.get(), smart_prefix, template, case_format, pad_len)

        preview_text = f"Original:  {dummy_orig}\nPreview:   {dummy_new}\n\n(Click anywhere to close)"

        if hasattr(self, 'preview_window') and self.preview_window.winfo_exists():
            self.preview_window.destroy()

        self.preview_window = tk.Toplevel(self)
        self.preview_window.wm_overrideredirect(True)
        x = self.winfo_rootx() + 20
        y = self.winfo_rooty() + self.winfo_height() - 250
        self.preview_window.wm_geometry(f"+{x}+{y}")
        frame = tk.Frame(self.preview_window, highlightbackground="#48484A", highlightthickness=1)
        frame.pack()

        lbl = tk.Label(frame, text=preview_text, justify='left', background="#2C2C2E", foreground="white", 
                       relief='flat', font=("Courier", 13, "normal"), padx=20, pady=15, wraplength=480, cursor="hand2")
        lbl.pack(ipadx=1)
        lbl.bind("<Button-1>", lambda e: self.preview_window.destroy())
        self.preview_window.after(8000, lambda: self.preview_window.destroy() if self.preview_window.winfo_exists() else None)

    def execute_rename(self):
        lang, cat, aspect, app, find, repl, smart_prefix, rename_folders, template, auto_asp, case_format, pad_str = self.format_inputs()
        if not all([cat, aspect, app, template]):
            self.set_status("Missing fields.", "#FF453A")
            return messagebox.showwarning("Missing Info", "Please fill out Category, Aspect Ratio, Application, and Template.")

        use_regex = self.regex_var.get()
        pad_len = len(pad_str)
        sort_method = self.sort_var.get()
        gen_md5 = self.md5_var.get()
        targets = []
        self.last_generated_logs = []
        self.last_generated_excel = ""
        
        if self.batch_mode_var.get():
            if not self.batch_folders: return self.set_status("Batch list is empty.", "#FF453A")
            for item in self.batch_folders: targets.append((item['path'], item['stem_var'].get()))
        else:
            p = filedialog.askdirectory()
            if p: targets.append((p, self.stem_var.get()))
            
        if not targets: return
        
        # --- PRE-FLIGHT SAFETY CHECKER ---
        self.set_status("Running Pre-Flight Safety Scan...", "#0A84FF")
        self.update()
        unsafe_files = []
        for base_folder, stem in targets:
            for root_dir, dirs, files in os.walk(base_folder):
                for f in files:
                    if f.startswith('.'): continue
                    if os.path.splitext(f)[1].lower() not in ALLOWED_EXTENSIONS: continue
                    dummy, _ = generate_safe_name(".", f, stem, lang, cat, aspect, app, find, repl, use_regex, smart_prefix, template, case_format, pad_len)
                    if dummy:
                        if len(dummy) > 150:
                            unsafe_files.append(f"{f} (Exceeds 150 chars)")
        
        if unsafe_files:
            msg = "PRE-FLIGHT ABORT: The following files will generate unsafe names that CMS databases reject:\n\n" + "\n".join(unsafe_files[:10])
            if len(unsafe_files) > 10: msg += f"\n...and {len(unsafe_files)-10} more."
            messagebox.showerror("Safety Scan Failed", msg)
            self.set_status("Aborted: Unsafe filenames detected.", "#FF453A")
            return

        total_files = sum([1 for f, _ in targets for r, _, fl in os.walk(f) for file in fl if not file.startswith('.') and os.path.splitext(file)[1].lower() in ALLOWED_EXTENSIONS])
        if total_files == 0:
            return self.set_status("No valid media files found in selected folders.", "#FF453A")
        
        self.progress_bar.pack(side="right", padx=10)
        self.progress_bar.set(0)
        self.config(cursor="watch")
        
        total_renamed = 0
        export_data = []
        
        stem_display_map = {
            "Complete": "Complete Video",
            "Video": "Video Only",
            "Speech": "Speech Only",
            "Music": "Music Only"
        }

        bundle_aspect_cache = {}
        if auto_asp:
            self.set_status("Analyzing bundle dimensions...", "#0A84FF")
            self.update()
            for base_folder, stem in targets:
                for root_dir, _, files in os.walk(base_folder):
                    for f in files:
                        if f.startswith('.'): continue
                        ext = os.path.splitext(f)[1].lower()
                        if ext not in ALLOWED_EXTENSIONS or ext in {'.wav', '.mp3', '.m4a', '.aac'}: continue
                        
                        name_id, _ = os.path.splitext(f)
                        replaced_id = apply_find_replace(name_id, find, repl, use_regex, smart_prefix)
                        pure_base = re.sub(r'[^A-Za-z0-9]', '', replaced_id).lower()
                        
                        if pure_base not in bundle_aspect_cache or stem == "Complete":
                            detected_aspect = get_auto_aspect_ratio(os.path.join(root_dir, f), aspect)
                            bundle_aspect_cache[pure_base] = detected_aspect

        try:
            for base_folder, stem in targets:
                content_type_str = stem_display_map.get(stem, "Unknown")
                
                for root_dir, dirs, files in os.walk(base_folder, topdown=True):
                    logs = []
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    
                    if rename_folders and find:
                        for i in range(len(dirs)):
                            old_d = dirs[i]
                            new_d = apply_find_replace(old_d, find, repl, use_regex, smart_prefix)
                            if new_d != old_d:
                                os.rename(os.path.join(root_dir, old_d), os.path.join(root_dir, new_d))
                                dirs[i] = new_d 
                                logs.append(f"DIR_RENAME\t{old_d}\t{new_d}\n")
                    
                    log_path = os.path.join(root_dir, f"Backup_{os.path.basename(root_dir)}_{timestamp}.txt")
                    valid_files = [f for f in files if not f.startswith('.') and os.path.isfile(os.path.join(root_dir, f))]
                    
                    # --- SMART SORTING ---
                    if sort_method == "Alphabetical":
                        valid_files.sort(key=lambda x: x.lower())
                    elif sort_method == "Creation Date":
                        valid_files.sort(key=lambda x: get_creation_time(os.path.join(root_dir, x)))
                    elif sort_method == "Modified Date":
                        valid_files.sort(key=lambda x: os.path.getmtime(os.path.join(root_dir, x)))
                    
                    for f in valid_files:
                        current_file_path = os.path.join(root_dir, f)
                        
                        name_id, _ = os.path.splitext(f)
                        replaced_id = apply_find_replace(name_id, find, repl, use_regex, smart_prefix)
                        pure_base = re.sub(r'[^A-Za-z0-9]', '', replaced_id).lower()
                        
                        final_aspect = bundle_aspect_cache.get(pure_base, aspect) if auto_asp else aspect
                        
                        new, final_name_id = generate_safe_name(root_dir, f, stem, lang, cat, final_aspect, app, find, repl, use_regex, smart_prefix, template, case_format, pad_len)
                        
                        if new:
                            new_full_path = os.path.join(root_dir, new)
                            os.rename(current_file_path, new_full_path)
                            logs.append(f"{f}\t{new}\n")
                            
                            duration_str = get_media_duration(new_full_path)
                            
                            md5_hash = "Skipped"
                            if gen_md5:
                                self.set_status(f"Generating MD5: {new}...", "#0A84FF")
                                self.update()
                                md5_hash = calculate_md5(new_full_path)
                                
                            export_data.append([root_dir, f, new, final_name_id, content_type_str, duration_str, md5_hash])
                            total_renamed += 1
                            
                        self.progress_bar.set(total_renamed / total_files)
                        self.set_status(f"Processed {total_renamed}/{total_files} files...")
                        self.update()

                    if logs:
                        with open(log_path, 'w') as logfile:
                            logfile.write("ORIGINAL_NAME\tNEW_NAME\n")
                            logfile.writelines(logs)
                        self.last_generated_logs.append(log_path)
            
            if export_data:
                export_data.sort(key=workout_bundle_sort_key)
                
                if self.batch_mode_var.get():
                    parent_dir = os.path.dirname(targets[0][0])
                    folder_name = os.path.basename(parent_dir)
                    save_dir = parent_dir
                else:
                    folder_name = os.path.basename(targets[0][0])
                    save_dir = targets[0][0]
                
                if not folder_name: folder_name = "Project"
                
                excel_path = os.path.join(save_dir, f"{folder_name}_Rename_Report.xlsx")
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Delivery Manifest"
                
                ws.append(["Original Name", "Content Type", "New Name", "Duration", "MD5 Checksum"])
                header_font = Font(bold=True)
                for col in ["A", "B", "C", "D", "E"]:
                    ws[f"{col}1"].font = header_font
                
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 60
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 35
                
                for row in export_data:
                    orig_full = row[1]
                    new_name = row[2]
                    content_type = row[4]
                    duration = row[5]
                    md5_hash = row[6]
                    
                    orig_base, _ = os.path.splitext(orig_full)
                    ws.append([orig_base, content_type, new_name, duration, md5_hash])
                    
                    if content_type == "Complete Video":
                        ws[f"A{ws.max_row}"].font = Font(bold=True)
                
                wb.save(excel_path)
                self.last_generated_excel = excel_path 
                self.set_status(f"Success: Renamed {total_renamed} files. Manifest saved.", "#34C759")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.set_status("Error occurred during rename.", "#FF453A")
            
        finally:
            self.config(cursor="")
            self.progress_bar.pack_forget()
            
            if self.batch_mode_var.get() and total_renamed > 0:
                self.batch_mode_var.set(False)
                self.toggle_batch_mode()
                for widget in self.scroll_list.winfo_children(): widget.destroy()
                self.batch_folders.clear()

    def execute_revert(self):
        if self.last_generated_logs:
            if messagebox.askyesno("Quick Undo", f"Would you like to instantly undo the previous rename operation?\n\nThis will safely revert {len(self.last_generated_logs)} directories from the bottom up."):
                self.config(cursor="watch")
                self.update()
                logs_to_revert = reversed(self.last_generated_logs)
                total_reverted = sum(process_revert_from_log(log) for log in logs_to_revert)
                if self.last_generated_excel and os.path.exists(self.last_generated_excel):
                    move_to_trash(self.last_generated_excel)
                self.last_generated_logs.clear()
                self.last_generated_excel = ""
                self.config(cursor="")
                self.set_status(f"Reverted files & folders safely. Logs moved to Trash.", "#34C759")
            return

        if self.batch_mode_var.get():
            if not self.batch_folders: return self.set_status("No folders in batch list to revert.", "#FF453A")
            logs_to_revert = []
            for item in self.batch_folders:
                base_folder = item['path']
                for root_dir, dirs, files in os.walk(base_folder, topdown=False):
                    logs = [f for f in files if f.startswith("Backup_") and f.endswith(".txt")]
                    if logs:
                        latest_log = max(logs, key=lambda x: os.path.getmtime(os.path.join(root_dir, x)))
                        logs_to_revert.append(os.path.join(root_dir, latest_log))
                        
            if not logs_to_revert: return self.set_status("No recent backup logs found.", "#FF453A")
            if messagebox.askyesno("Revert Batch", f"Found {len(logs_to_revert)} backup logs across subfolders.\nAuto-revert all and delete associated report?"):
                self.config(cursor="watch")
                self.update()
                total_reverted = sum(process_revert_from_log(log) for log in logs_to_revert)
                parent_dir = os.path.dirname(self.batch_folders[0]['path'])
                folder_name = os.path.basename(parent_dir)
                if not folder_name: folder_name = "Project"
                possible_excel = os.path.join(parent_dir, f"{folder_name}_Rename_Report.xlsx")
                if os.path.exists(possible_excel): move_to_trash(possible_excel)
                self.config(cursor="")
                self.set_status(f"Reverted files & folders safely. Logs moved to Trash.", "#34C759")
                
        else:
            log_file = filedialog.askopenfilename(title="Select Backup Log File", filetypes=[("Text Files", "*.txt")])
            if log_file: 
                reverted = process_revert_from_log(log_file)
                folder_dir = os.path.dirname(log_file)
                folder_name = os.path.basename(folder_dir)
                possible_excel = os.path.join(folder_dir, f"{folder_name}_Rename_Report.xlsx")
                if os.path.exists(possible_excel): move_to_trash(possible_excel)
                self.set_status(f"Reverted safely. Report moved to Trash.", "#34C759")

    def execute_delete_logs(self):
        warning_msg = "WARNING: Deleting backup logs means you can no longer use the Undo button for those files.\n\nDo you want to proceed?"
        if not messagebox.askyesno("Deep Sweep Backup Logs", warning_msg, icon='warning'): return
        folder = filedialog.askdirectory(title="Select the Master Folder to Deep Clean")
        if not folder: return
        self.config(cursor="watch")
        self.update()
        logs_deleted = 0
        for root_dir, dirs, files in os.walk(folder):
            for f in files:
                if f.startswith("Backup_") and f.endswith(".txt"):
                    move_to_trash(os.path.join(root_dir, f))
                    logs_deleted += 1
        self.config(cursor="")
        if logs_deleted > 0:
            self.set_status(f"Deep Sweep: Cleaned {logs_deleted} log files.", "#34C759")
            self.last_generated_logs.clear() 
        else:
            self.set_status("No backup logs found in that folder.", "#8E8E93")

    # --- Persistent Settings ---
    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r') as f: d = json.load(f)
                
                if 'presets' in d:
                    self.presets = d['presets']
                else:
                    self.presets = {"Default": {}}
                    
                last_prof = d.get('last_profile', 'Default')
                if last_prof in self.presets:
                    self.preset_var.set(last_prof)
                    self.preset_cb.configure(values=list(self.presets.keys()))
                    self.load_preset_ui(last_prof)
                else:
                    self.lang_var.set(d.get('l', 'en'))
                    self.category_var.set(d.get('c', 'stretching'))
                    self.app_var.set(d.get('a', 'yg'))
                    self.aspect_var.set(d.get('asp', '16x9'))
                    self.auto_aspect_var.set(d.get('auto_asp', True))
                    self.stem_var.set(d.get('stem_single', 'Complete'))
                    self.smart_prefix_var.set(d.get('smart', False))
                    self.rename_folders_var.set(d.get('rn_fld', False))
                    self.padding_var.set(d.get('pad', '01'))
                    self.sort_var.set(d.get('sort', 'Alphabetical'))
                    self.md5_var.set(d.get('md5', False))
                    
                    tmpl_val = d.get('tmpl', '')
                    if not tmpl_val:
                        tmpl_val = "{name}{stem}+lang={lang}&category={cat}&aspect={aspect}&app={app}&master"
                    self.template_textbox.delete("1.0", "end")
                    self.template_textbox.insert("1.0", tmpl_val)
                    self.highlight_template()
                        
                self.toggle_batch_mode()
            except: pass
        else:
            self.preset_cb.configure(values=list(self.presets.keys()))

    def save_settings_and_close(self):
        current_prof = self.preset_var.get()
        if current_prof in self.presets:
            self.presets[current_prof] = {
                'l': self.lang_var.get(), 'c': self.category_var.get(), 'a': self.app_var.get(),
                'asp': self.aspect_var.get(), 'auto_asp': self.auto_aspect_var.get(),
                'find': self.find_var.get(), 'repl': self.replace_var.get(),
                'regex': self.regex_var.get(), 'smart': self.smart_prefix_var.get(),
                'rn_fld': self.rename_folders_var.get(), 'case': self.case_var.get(),
                'pad': self.padding_var.get(), 'sort': self.sort_var.get(), 'md5': self.md5_var.get(),
                'tmpl': self.template_textbox.get("1.0", "end-1c").strip()
            }
            
        d = {
            'presets': self.presets,
            'last_profile': current_prof,
            'stem_single': self.stem_var.get()
        }
        with open(SETTINGS_FILE, 'w') as f: json.dump(d, f)
        self.destroy()

if __name__ == "__main__":
    app = RenamerApp()
    app.mainloop()